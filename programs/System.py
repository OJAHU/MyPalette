from contextlib import redirect_stdout
from datetime import datetime, date
import os
from win11toast import toast
import win32com.client
import openpyxl
import base64

from email import message_from_bytes
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from email.header import decode_header
from google.auth.transport.requests import Request
from google.auth.exceptions import RefreshError

'''
・挨拶をする通知
・別プログラムMeMOを呼び出し、そのうちのVBAプログラムを実行
・MeMO内の今日まで・明日までのタスクのみを抽出し、記録
・ユーザにMeMOを開くかどうかを通知
・MeMO内の時間割の取得
・Gmail APIでメールの取得
'''

#通知システム
class System(): 
    def __init__(self, MeMO: str,
                 task: str,
                 vba: str,
                 token: str,
                 credentials: str):

        self.MeMO = MeMO #Excelファイルのパス
        self.task = task #タスク管理テキストファイルのパス
        self.vba = vba #実行するVBA名
        self.token = token #Gmail APIのtoken
        self.credentials = credentials #Gmail APIのcrendentails
        
        # コンソールにテキスト非表示
        with redirect_stdout(open(os.devnull, "w")):
            try:
                self.greeting()
                self.call_vba()
                self.read_task()
                self.open_MeMO()
                self.get_timetable()
                self.get_mail()
            except Exception as e:
                toast("エラー",
                      f"エラーが発生しました：{e}")
    
    def greeting(self):
        #挨拶通知
        
        now = datetime.now() #現在時刻の取得
        greet = ""
        if 4 <= now.hour < 12:
            greet = "おはようございます！"
        elif 12 <= now.hour < 18:
            greet = "こんにちは！"
        else:
            greet = "こんばんは！"
        
        #通知
        toast(greet,
              f"現在は{now.strftime('%Y年%m月%d日 %H:%M')}です。",
              duration = "short")
    
    def call_vba(self): #エクセルのVBA呼び出し
        excel = win32com.client.Dispatch("Excel.Application") #Excelを起動
        excel.Visible = 0 #Excelを非表示で実行
        excel.Workbooks.Open(self.MeMO) #指定ファイルを開く
        excel.Application.Run(self.vba) #VBA実行
        excel.Workbooks(1).Close(SaveChanges=1) #保存して閉じる
        excel.Application.Quit() #Excelを終了

    def read_task(self):
        self.until_today = 0 #今日までの課題数
        self.until_tommorow = 0 #明日までの課題数
        self.tasks = [] #タスク格納リスト

        with open(self.task, "r", encoding="utf-8") as f:
            task_lst = f.read().replace("\n", ",").split("BOT")
            del task_lst[-1] #余白を削除

            for task in task_lst:
                task = [i.replace("\ufeff","") for i in task.split(",") if i != ""]
                self.tasks.append(task)

                if task[0] == "0":
                    self.until_today += 1
                elif task[0] == "1":
                    self.until_tommorow += 1
        
        #通知表示
        toast("今日までのタスク",
              (f"今日までのタスクは{self.until_today}件あります！" if self.until_today > 0 else "今日までのタスクはありません"))
        toast("明日までのタスク",
              (f"明日までのタスクは{self.until_tommorow}件あります！" if self.until_tommorow > 0 else "明日までのタスクはありません"))
    
    def open_MeMO(self):
        #通知を押すとMeMOが開く
        toast("MeMOを開くにはクリックをしてください",
              on_click=self.MeMO, 
              duration="short")
    
    def get_timetable(self):
        #今日の時間割の取得
        date_num = date.today().weekday()
        self.subject_lst = []
        self.time_lst = []
        
        if date_num < 5:
            book = openpyxl.load_workbook(self.MeMO)
            
            sheet = book["時間割"]        
            
            cell_y = 3
            cell_x = 3 + date_num
            
            for _ in range(6):                
                if cell_y != 5:   
                    subject = sheet.cell(cell_y, cell_x).value
                    if subject == None:
                        self.subject_lst.append("なし")
                    else:
                        self.subject_lst.append(subject)
                        
                    time = sheet.cell(cell_y, 2).value
                    self.time_lst.append(time)

                cell_y += 1

    def get_mail(self):
        #Gmail APIを使って受信メールの取得
        
        SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
        
        #トークンファイルから資格情報を読み込み
        creds = Credentials.from_authorized_user_file(self.token, SCOPES)

        #認証がない、期限切れの場合は再認証
        if not creds or not creds.valid:
            try:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                else:
                    flow = InstalledAppFlow.from_client_secrets_file(self.credentials, SCOPES)
                    creds = flow.run_local_server(port=0)
                
            except RefreshError:
                if os.path.exists(self.token):
                    os.remove(self.token)
                    
                flow = InstalledAppFlow.from_client_secrets_file(self.credentials, SCOPES)
                creds = flow.run_local_server(port=0)
            
            with open(self.token, "w") as token:
                token.write(creds.to_json())
                
        #Gmail APIのサービスオブジェクトを構築
        service = build("gmail", "v1", credentials=creds)
        #受信メールのメッセージのID一覧（10件）
        results = service.users().messages().list(userId="me", maxResults=10).execute()
        messages = results.get("messages", [])
        
        self.msgs = []
        if not messages: #メールがなければ
            self.msgs.append({"Subject": "メールが見つかりません", "From": "Not Found"})
        else: #メールがあるなら
            for msg in messages:
                msg_id = msg["id"]
                
                #メールの詳細をRAW形式で取得
                msg_data = service.users().messages().get(userId="me", id=msg_id, format="raw").execute()
                #デコード
                raw_msg = base64.urlsafe_b64decode(msg_data["raw"].encode("UTF-8"))
                #MIME形式のメールオブジェクトに変換
                mime_msg = message_from_bytes(raw_msg)

                #ヘッダーの文字コードをデコードする関数
                def decode_words(st: str) -> str:
                    decoded = decode_header(st)

                    text = "".join([
                        fragment.decode(encoding or 'utf-8') if isinstance(fragment, bytes) else fragment
                        for fragment, encoding in decoded
                    ])

                    return text
                
                #件名と差出人を取得してそれぞれデコード
                subject = decode_words(mime_msg.get("Subject", "(件名なし)"))
                from_name = decode_words(mime_msg.get("From", "(差出人不明)"))
                
                self.msgs.append({"Subject": subject, "From": from_name})

